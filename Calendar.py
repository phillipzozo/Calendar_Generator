import time
import datetime
from datetime import timedelta
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, PatternFill, Side, Border, Font
from openpyxl.utils import get_column_letter


# 設定年份
year = int(input('year: '))
monthly_writer_path = f"{year}_calendar.xlsx"

# 找出該年份的第一天與最後一天
start_date = datetime.date(year, 1, 1)
end_date = datetime.date(year, 12, 31)

# 找到第一個週日作為起點
while start_date.weekday() != 6:  # Sunday is 6
    start_date -= timedelta(days=1)

# 建立週為單位的表格，每列為一週
rows = []
week_days = ["Sunday", "Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Saturday"]
current_date = start_date

# 判斷有無跨月
new_month = True
new_month_days = 0
while current_date <= end_date:
    # 換月但顯示下個月的同週
    week_end_date = current_date + timedelta(days=6)
    this_month = current_date.month
    previous_week_start_date = current_date - timedelta(days=7)
    previous_week_end_date = previous_week_start_date + timedelta(days=6)
    # 每月開頭有一列月份
    #print(f'week_end_date == this_month')
    #print(f'week_end_date == this_month')
    if week_end_date.month == this_month \
        and previous_week_start_date.month == previous_week_end_date.month \
        and week_end_date.month != previous_week_end_date.month:
        # 如果本月週開頭與結尾同月且上月週開頭與結尾同月且本月週與上月週不同月
        month_row = {}
        month_row["Week"] = 'Month'
        for i, day in enumerate(week_days):
            if i == 0:
                month_row[day] = f'{week_end_date.month}'
            else:
                month_row[day] = ''
        rows.append(month_row)
        
    elif new_month == True:
        # 如果本月週開頭與結尾不同月
        month_row = {}
        month_row["Week"] = 'Month'
        for i, day in enumerate(week_days):
            if i == 0:
                month_row[day] = f'{week_end_date.month}'
            else:
                month_row[day] = ''
        rows.append(month_row)
        
        # 如果本週最後一天不同月則 this_month 變成下個月
        if week_end_date.month != this_month:
            this_month = week_end_date.month
        # 讓下一個迴圈會印出同一週，並且避開前一年12月
        if week_end_date.month != 1:
            current_date -= timedelta(days=7)
        
        new_month = False
    
    week = {}
    week_start = current_date
    iso_week = current_date.isocalendar()[1]
    week["Week"] = f"Week {iso_week}"
    for i, day in enumerate(week_days):
        date_str = ""
        current_day = current_date + timedelta(days=i)
        # 僅包含今年的資料
        if current_day.year == year:
            if current_day.month == this_month:
                date_str = f"{current_day.day}"
                new_month = False
            else:
                date_str = ''
                new_month = True
            
        week[day] = date_str
    rows.append(week)
    for _ in range(4):
        week = {}
        week["Week"] = ''
        for i, day in enumerate(week_days):
            week[day] = ''
        rows.append(week)
    current_date += timedelta(days=7)

# 建立 Excel 檔案，每月一個工作表
df = pd.DataFrame(rows)
df.to_excel(monthly_writer_path, sheet_name=f'{year}', index=False)

# 標題列格式
titleAlignment = Alignment(horizontal='center', vertical='center')
titleFont = Font(name='Calibri', size=13, bold=True, color='FFFFFF')
titleFill = PatternFill(start_color="46556e", end_color="46556e", fill_type="solid")

# 月份列格式
monthAlignment = Alignment(horizontal='center', vertical='center')
monthFont = Font(name='Calibri', color='FFFFFF')
monthFill = PatternFill(start_color="7a889f", end_color="7a889f", fill_type="solid")

# 內容格式
wb = load_workbook(monthly_writer_path)
#'dashed', 'medium', 'thick', 'thin', 'dashDotDot', 'mediumDashDotDot', 'mediumDashed', 
#'double', 'mediumDashDot', 'dotted', 'slantDashDot', 'hair', 'dashDot'
outterSide = Side(color="508ab8", border_style="thick")
innerSide = Side(color="bad9f5", border_style="medium")
writeSide = Side(color="ffffff", border_style="medium")
innerFillSide = Side(color="d8e8eb", border_style="medium")
weekwndFill = PatternFill(start_color="d8e8eb", end_color="bcd9dd", fill_type="solid")

for sheet in wb.sheetnames:
    ws = wb[sheet]
    max_col = ws.max_column
    max_row = ws.max_row
    
    ws_iter_rows = list(ws.iter_rows(min_row=1, max_row=max_row, min_col=1, max_col=8))
    ws_iter_rows_len = len(ws_iter_rows)
    
    # 設定標題列格式
    ws.row_dimensions[1].height = 30
    # 設定week行格式
    ws.column_dimensions['A'].width = 10
    # 設定星期行寬度格式
    for i in ['B', 'C', 'D', 'E', 'F', 'G', 'H']:
        ws.column_dimensions[i].width = 15
    # 加邊框及填充顏色
    for i, row in enumerate(ws_iter_rows):
        for j, cell in enumerate(row):
            # 'Week'為標題列，'Week '為內容
            row_0_value = str(row[0].value)
            #print(f'row[0].value {row[0].value}', end=' ')
            border_left = innerSide
            border_right = innerSide
            border_top = innerSide
            border_bottom = innerSide
            fill = weekwndFill
            
            if 'Week' in row_0_value \
                or 'Month' in row_0_value:
                # 如果是日期列或月份列，邊框都有，但week行下邊框空白
                if 'Month' in row_0_value and 'Month' not in str(cell.value):
                    month_cells = f'{row[1].column_letter}{row[1].row}:{row[-1].column_letter}{row[-1].row}'
                    ws.merge_cells(month_cells)
                    cell.alignment = monthAlignment
                    cell.font = monthFont
                    cell.fill = monthFill
                    
                    border_right = outterSide
                    fill = None
                elif 'Week ' in str(cell.value):
                    #print('下邊框空白')
                    border_bottom = writeSide
                else:
                    #print('邊框都有')
                    fill = None
                    pass
            elif i - 1 >= 0 \
                and ws_iter_rows[i-1][0].value != None \
                and 'Week ' not in row_0_value \
                and 'Week ' not in str(ws_iter_rows[i-1][j].value):
                # 如果上列不是空白且不是日期列，下邊框空白，且上列不是week行
                #print('下邊框空白')
                border_bottom = writeSide
            elif ws_iter_rows_len < i + 1 \
                and ws_iter_rows[i+1][j].value != None \
                and 'Week ' not in row_0_value:
                # 如果下列不是空白且不是日期列，上邊框空白
                #print('上邊框空白')
                border_top = writeSide
            else:
                # 上下邊框空白
                #print('上下邊框空白')
                border_top = writeSide
                border_bottom = writeSide
                # 假日邊框填充背景顏色
                if not (i == 0 or j != 1 and j != len(row)-1):
                    border_top = innerFillSide
                    border_bottom = innerFillSide
                
            # 整體外邊框
            if j == 0:
                border_left = outterSide
            if j == len(row)-1:
                border_right = outterSide
            if i == 0:
                border_top = outterSide
            if i == ws_iter_rows_len-1:
                border_bottom = outterSide
                
            # 假日填充背景
            if i == 0:
                cell.alignment = titleAlignment
                cell.font = titleFont
                fill = titleFill
            elif j != 1 and j != len(row)-1:
                fill = None
                
            if fill != None:
                cell.fill = fill
                
            cell.border = Border(left=border_left, right=border_right, top=border_top, bottom=border_bottom)
            

wb.save(monthly_writer_path)
monthly_writer_path
